"""
Reading the data from cloth.xlsx and copying it into empty.xlsx
"""
import openpyxl

pathExcel = "C:\\Users\ShilpiDas\PycharmProjects\PythonAutomation\Selenium\cloth.xlsx"
workbook = openpyxl.load_workbook(pathExcel)

pathExcel_copy = "C:\\Users\ShilpiDas\PycharmProjects\PythonAutomation\Selenium\empty.xlsx"
workbook_copy = openpyxl.load_workbook(pathExcel)

sheet = workbook.active
sheet_copy = workbook_copy.active

# sheet=workbook.get_sheet_by_name("Sheet1")

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value, end="  ")
        sheet_copy.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value
    print()
workbook_copy.save(pathExcel_copy)
